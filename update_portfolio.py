import os
import shutil
import ast
import html
import math
import operator
import re
import tempfile
import urllib.error
import urllib.request
from datetime import datetime, timedelta
from html.parser import HTMLParser
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Image, Table, TableStyle, KeepTogether
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email import encoders
    HAS_EMAIL = True
except ImportError:
    HAS_EMAIL = False

try:
    from dotenv import load_dotenv
    HAS_DOTENV = True
    DOTENV_PATH = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.exists(DOTENV_PATH):
        load_dotenv(DOTENV_PATH)
    else:
        print("Warning: .env file not found; using system environment variables only.")
except ImportError:
    HAS_DOTENV = False
    print("Warning: python-dotenv not installed - .env file will not be loaded.")

try:
    from PyPDF2 import PdfWriter
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

CELL_REF_RE = re.compile(r"(?<![A-Za-z0-9_])(\$?[A-Z]{1,3}\$?\d+)(?![A-Za-z0-9_])")
SUMMARY_START_ROW = 2
ASSET_TABLE_START_ROW = 8

# PDF and Email Configuration
if HAS_REPORTLAB:
    PDF_PAGE_SIZE = letter
    PDF_LEFT_MARGIN = 0.5
    PDF_RIGHT_MARGIN = 0.5
else:
    PDF_PAGE_SIZE = None
    PDF_LEFT_MARGIN = 0.5
    PDF_RIGHT_MARGIN = 0.5

PDF_TITLE_SIZE = 20
PDF_HEADING_SIZE = 14
PDF_TEXT_SIZE = 10
MONTHS_TO_DISPLAY = 6

EMAIL_SMTP_SERVER = os.environ.get("EMAIL_SMTP_SERVER", "smtp.gmail.com")
EMAIL_SMTP_PORT = int(os.environ.get("EMAIL_SMTP_PORT", "587"))
EMAIL_SMTP_USERNAME = os.environ.get("EMAIL_SMTP_USERNAME", "")
EMAIL_SMTP_PASSWORD = os.environ.get("EMAIL_SMTP_PASSWORD", "")
EMAIL_FROM = os.environ.get("EMAIL_FROM", EMAIL_SMTP_USERNAME)
EMAIL_USE_TLS = os.environ.get("EMAIL_USE_TLS", "True").lower() in ("true", "1", "yes")
EMAIL_USE_SSL = os.environ.get("EMAIL_USE_SSL", "False").lower() in ("true", "1", "yes")

SECTOR_WEIGHT_LIMIT = 20
SINGLE_ASSET_WEIGHT_LIMIT = 10
SECTOR_RISK_EXCLUSIONS = {"Cash", "Fixed Income"}
TRANSACTIONS_SHEET = "Transactions"
STATE_SHEET = "Portfolio_State"
DIVIDEND_INPUT_SHEET = "Dividend Tracking"
DIVIDEND_OUTPUT_SHEET = "Dividends"
PERFORMANCE_HISTORY_SHEET = "Performance History"
NSE_PRICE_SOURCE_URL = "https://www.mansamarkets.com/kenya"
NSE_PRICE_SOURCE_NAME = "NSE live market data"
NSE_PRICE_REFRESHED_AT = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
NSE_ASSET_TICKERS = {
    "Co-op Bank": "COOP",
    "Equity": "EQTY",
    "KCB": "KCB",
    "NCBA": "NCBA",
    "Safaricom": "SCOM",
    "Jubilee": "JUB",
    "CIC": "CIC",
    "Britam": "BRIT",
    "KenGen": "KEGN",
    "Kenya Power": "KPLC",
    "Total Energies Marketing": "TOTL",
}
SECTOR_LABELS = {
    "Banking": "Banking",
    "Telecommunication": "Telecom",
    "Insurance": "Insurance",
    "Energy": "Energy",
    "ETFs": "ETF",
    "Fixed Income": "Fixed Income",
    "Liquid Cash": "Cash",
}

DARK_OLIVE = "3B4436"
CREAM = "F1E9CB"
TEXT_DARK = "2F332E"
WARM_BEIGE = "E6DFD3"
WARM_WHITE = "FDFBF7"
SOFT_BEIGE = "EAECE6"
BORDER_COLOR = "B8AA91"
TRANSACTION_BASE_COLUMNS = [
    "Date",
    "Asset",
    "Action",
    "Quantity",
    "Price",
    "Broker",
    "Fees",
    "Benefits",
]
TRANSACTION_CALC_COLUMNS = [
    "Realized Gain",
    "Unrealized Gain",
    "Cost Basis",
    "Average Purchase Price",
    "Position Quantity",
]
TRANSACTION_COLUMNS = TRANSACTION_BASE_COLUMNS + TRANSACTION_CALC_COLUMNS
STATE_COLUMNS = [
    "Asset",
    "Sector",
    "Shares",
    "Buy Price",
    "Current Price",
    "Market Value",
]
NSE_PRICE_TABLE_HEADERS = ["Asset", "Ticker", "NSE Price", "Price Source", "Last Updated"]

ALLOWED_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.Pow: operator.pow,
}

ALLOWED_UNARY_OPERATORS = {
    ast.UAdd: operator.pos,
    ast.USub: operator.neg,
}


def evaluate_math_node(node):
    if isinstance(node, ast.Expression):
        return evaluate_math_node(node.body)

    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return node.value

    if isinstance(node, ast.BinOp) and type(node.op) in ALLOWED_OPERATORS:
        return ALLOWED_OPERATORS[type(node.op)](
            evaluate_math_node(node.left),
            evaluate_math_node(node.right),
        )

    if isinstance(node, ast.UnaryOp) and type(node.op) in ALLOWED_UNARY_OPERATORS:
        return ALLOWED_UNARY_OPERATORS[type(node.op)](evaluate_math_node(node.operand))

    raise ValueError("Unsupported formula expression")


def numeric_cell_value(worksheet, coordinate, seen=None):
    coordinate = coordinate.replace("$", "")
    seen = seen or set()

    if coordinate in seen:
        raise ValueError(f"Circular formula reference at {coordinate}")

    seen.add(coordinate)
    value = worksheet[coordinate].value

    if isinstance(value, str) and value.startswith("="):
        return evaluate_excel_formula(worksheet, value, seen)

    return pd.to_numeric(value, errors="coerce")


def evaluate_excel_formula(worksheet, formula, seen=None):
    expression = formula.lstrip("=").replace("^", "**")

    def replace_cell_reference(match):
        value = numeric_cell_value(worksheet, match.group(1), seen)
        if pd.isna(value):
            raise ValueError("Formula references a blank or non-numeric cell")
        return str(float(value))

    expression = CELL_REF_RE.sub(replace_cell_reference, expression)

    if re.search(r"[A-Za-z]", expression):
        raise ValueError("Unsupported formula expression")

    return evaluate_math_node(ast.parse(expression, mode="eval"))


def formula_backed_column_values(file_path, sheet_name, header_idx, column_index, row_count):
    workbook = load_workbook(file_path, data_only=False)
    worksheet = workbook[sheet_name]
    values = []

    for row_idx in range(row_count):
        excel_row = header_idx + 2 + row_idx
        cell = worksheet.cell(row=excel_row, column=column_index)

        try:
            value = (
                evaluate_excel_formula(worksheet, cell.value)
                if isinstance(cell.value, str) and cell.value.startswith("=")
                else cell.value
            )
        except ValueError:
            value = None

        values.append(value)

    return pd.Series(pd.to_numeric(values, errors="coerce"))


class TextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []

    def handle_data(self, data):
        if data:
            self.parts.append(data)

    def get_text(self):
        return html.unescape(" ".join(self.parts))


def fetch_url_text(url):
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        },
    )

    with urllib.request.urlopen(request, timeout=20) as response:
        return response.read().decode("utf-8", errors="replace")


def extract_current_price_from_html(page_html):
    extractor = TextExtractor()
    extractor.feed(page_html)
    page_text = re.sub(r"\s+", " ", extractor.get_text()).strip()

    match = re.search(r"Current Price\s*KSh\s*([\d,]+(?:\.\d+)?)", page_text, re.IGNORECASE)
    if match:
        return float(match.group(1).replace(",", ""))

    match = re.search(r"KSh\s*([\d,]+(?:\.\d+)?)\s*Change", page_text, re.IGNORECASE)
    if match:
        return float(match.group(1).replace(",", ""))

    return None


def fetch_nse_price_for_ticker(ticker):
    url = f"https://www.mansamarkets.com/kenya/{ticker.lower()}"
    try:
        page_html = fetch_url_text(url)
        price = extract_current_price_from_html(page_html)
        if price is None:
            raise ValueError(f"Could not parse NSE price from {url}")
        return {
            "ticker": ticker,
            "price": price,
            "source": NSE_PRICE_SOURCE_NAME,
            "updated": NSE_PRICE_REFRESHED_AT,
            "url": url,
        }
    except Exception:
        return {
            "ticker": ticker,
            "price": None,
            "source": f"{NSE_PRICE_SOURCE_NAME} unavailable",
            "updated": NSE_PRICE_REFRESHED_AT,
            "url": url,
        }


def fetch_nse_prices_for_assets(asset_names):
    results = {}
    for asset_name in asset_names:
        ticker = NSE_ASSET_TICKERS.get(asset_name)
        if not ticker:
            continue
        results[asset_name] = fetch_nse_price_for_ticker(ticker)
    return results


def apply_nse_prices(holdings, nse_prices):
    holdings = holdings.copy()
    holdings["Ticker"] = holdings["Asset"].map(NSE_ASSET_TICKERS).fillna("")
    holdings["NSE Price"] = None
    holdings["Price Source"] = "Manual / non-NSE"
    holdings["Price Last Updated"] = ""

    for row_index, row in holdings.iterrows():
        asset_name = row["Asset"]
        price_data = nse_prices.get(asset_name)

        if not price_data:
            continue

        holdings.at[row_index, "Price Source"] = price_data["source"]
        holdings.at[row_index, "Price Last Updated"] = price_data["updated"]

        if price_data["price"] is not None:
            holdings.at[row_index, "Current Price"] = price_data["price"]
            holdings.at[row_index, "NSE Price"] = price_data["price"]

    return holdings


def build_nse_price_table(holdings):
    table = holdings[
        [
            "Asset",
            "Ticker",
            "Current Price",
            "Price Source",
            "Price Last Updated",
        ]
    ].copy()
    table.columns = NSE_PRICE_TABLE_HEADERS
    return table


def inject_nse_live_prices(holdings):
    asset_names = holdings["Asset"].dropna().unique()
    nse_prices = fetch_nse_prices_for_assets(asset_names)
    holdings = apply_nse_prices(holdings, nse_prices)
    return holdings


def empty_transactions():
    return pd.DataFrame(columns=TRANSACTION_COLUMNS)


def read_previous_sheet(file_path, sheet_name, columns):
    if not os.path.exists(file_path):
        return pd.DataFrame(columns=columns)

    try:
        workbook = pd.ExcelFile(file_path)
        if sheet_name not in workbook.sheet_names:
            return pd.DataFrame(columns=columns)

        data = pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame(columns=columns)

    for column in columns:
        if column not in data.columns:
            data[column] = None

    return data[columns].copy()


def build_current_state(holdings):
    if "Buy Price" not in holdings.columns:
        holdings = holdings.copy()
        holdings["Buy Price"] = holdings["Current Price"]

    state = holdings[STATE_COLUMNS].copy()
    state["Shares"] = pd.to_numeric(state["Shares"], errors="coerce").fillna(0)
    state["Current Price"] = pd.to_numeric(
        state["Current Price"],
        errors="coerce",
    ).fillna(0)
    state["Buy Price"] = pd.to_numeric(
        state["Buy Price"],
        errors="coerce",
    ).fillna(state["Current Price"])
    state["Market Value"] = pd.to_numeric(
        state["Market Value"],
        errors="coerce",
    ).fillna(0)
    return state


def detect_share_transactions(previous_state, current_state):
    if previous_state.empty:
        return pd.DataFrame(columns=TRANSACTION_BASE_COLUMNS)

    previous_state = previous_state.set_index("Asset", drop=False)
    current_state = current_state.set_index("Asset", drop=False)
    transaction_rows = []
    transaction_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for asset in sorted(set(previous_state.index).union(current_state.index)):
        previous_row = previous_state.loc[asset] if asset in previous_state.index else None
        current_row = current_state.loc[asset] if asset in current_state.index else None

        previous_shares = (
            pd.to_numeric(previous_row["Shares"], errors="coerce")
            if previous_row is not None
            else 0
        )
        current_shares = (
            pd.to_numeric(current_row["Shares"], errors="coerce")
            if current_row is not None
            else 0
        )
        previous_shares = 0 if pd.isna(previous_shares) else previous_shares
        current_shares = 0 if pd.isna(current_shares) else current_shares
        share_change = current_shares - previous_shares

        if abs(share_change) < 0.000001:
            continue

        price_source = current_row if current_row is not None else previous_row
        price = pd.to_numeric(price_source["Current Price"], errors="coerce")
        price = 0 if pd.isna(price) else price

        transaction_rows.append({
            "Date": transaction_date,
            "Asset": asset,
            "Action": "BUY" if share_change > 0 else "SELL",
            "Quantity": abs(share_change),
            "Price": price,
            "Broker": "",
            "Fees": 0,
            "Benefits": "",
        })

    return pd.DataFrame(transaction_rows, columns=TRANSACTION_BASE_COLUMNS)


def build_opening_transactions(state):
    if state.empty:
        return pd.DataFrame(columns=TRANSACTION_BASE_COLUMNS)

    opening_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    transaction_rows = []

    for _, row in state.iterrows():
        shares = pd.to_numeric(row["Shares"], errors="coerce")
        buy_price = pd.to_numeric(row.get("Buy Price"), errors="coerce")
        current_price = pd.to_numeric(row["Current Price"], errors="coerce")

        shares = 0 if pd.isna(shares) else shares
        buy_price = current_price if pd.isna(buy_price) else buy_price
        buy_price = 0 if pd.isna(buy_price) else buy_price

        if shares <= 0:
            continue

        transaction_rows.append({
            "Date": opening_date,
            "Asset": row["Asset"],
            "Action": "OPENING",
            "Quantity": shares,
            "Price": buy_price,
            "Broker": "",
            "Fees": 0,
            "Benefits": "Opening balance",
        })

    return pd.DataFrame(transaction_rows, columns=TRANSACTION_BASE_COLUMNS)


def calculate_transaction_metrics(transactions, current_state):
    if transactions.empty:
        return empty_transactions()

    transactions = transactions.copy()

    for column in TRANSACTION_BASE_COLUMNS:
        if column not in transactions.columns:
            transactions[column] = "" if column in ["Date", "Asset", "Action", "Broker", "Benefits"] else 0

    current_prices = current_state.set_index("Asset")["Current Price"].to_dict()
    positions = {}
    cost_bases = {}
    calculated_rows = []

    for _, transaction in transactions[TRANSACTION_BASE_COLUMNS].iterrows():
        asset = transaction["Asset"]
        action = str(transaction["Action"]).upper()
        quantity = pd.to_numeric(transaction["Quantity"], errors="coerce")
        price = pd.to_numeric(transaction["Price"], errors="coerce")
        fees = pd.to_numeric(transaction["Fees"], errors="coerce")
        quantity = 0 if pd.isna(quantity) else quantity
        price = 0 if pd.isna(price) else price
        fees = 0 if pd.isna(fees) else fees

        position_quantity = positions.get(asset, 0)
        cost_basis = cost_bases.get(asset, 0)
        average_price = cost_basis / position_quantity if position_quantity else 0
        realized_gain = 0

        if action in ["BUY", "OPENING"]:
            position_quantity += quantity
            cost_basis += (quantity * price) + fees
        elif action == "SELL":
            cost_removed = average_price * min(quantity, position_quantity)
            realized_gain = (quantity * price) - fees - cost_removed
            position_quantity -= quantity
            cost_basis = max(0, cost_basis - cost_removed)

        average_price = cost_basis / position_quantity if position_quantity else 0
        current_price = current_prices.get(asset, price)
        unrealized_gain = (position_quantity * current_price) - cost_basis

        positions[asset] = position_quantity
        cost_bases[asset] = cost_basis

        row = transaction.to_dict()
        row.update({
            "Realized Gain": realized_gain,
            "Unrealized Gain": unrealized_gain,
            "Cost Basis": cost_basis,
            "Average Purchase Price": average_price,
            "Position Quantity": position_quantity,
        })
        calculated_rows.append(row)

    return pd.DataFrame(calculated_rows, columns=TRANSACTION_COLUMNS)


def risk_label_from_score(score):
    if score >= 85:
        return "Low"
    if score >= 70:
        return "Medium"
    return "High"


def format_money(value):
    return f"KES {value:,.2f}"


def add_dynamic_asset_limits(holdings):
    holdings = holdings.copy()
    sector_asset_counts = holdings.groupby("Sector")["Asset"].transform("count")
    holdings["Asset Limit %"] = (
        SECTOR_WEIGHT_LIMIT / sector_asset_counts
    ).clip(upper=SINGLE_ASSET_WEIGHT_LIMIT)
    return holdings


def sell_amount_to_limit(current_value, total_value, limit_percent):
    limit = limit_percent / 100

    if limit >= 1:
        return 0

    return max(0, (current_value - (limit * total_value)) / (1 - limit))


def buy_amount_to_limit(current_value, total_value, limit_percent):
    limit = limit_percent / 100

    if limit <= 0:
        return 0

    return max(0, (current_value / limit) - total_value)


def build_rebalance_plan(holdings, total_value, sector_risk_total):
    risk_pool = holdings[~holdings["Sector"].isin(SECTOR_RISK_EXCLUSIONS)].copy()
    risk_pool = add_dynamic_asset_limits(risk_pool)

    sector_cap_value = (SECTOR_WEIGHT_LIMIT / 100) * total_value
    targets = risk_pool.set_index("Asset")["Market Value"].to_dict()
    asset_caps = (
        risk_pool.set_index("Asset")["Asset Limit %"] / 100 * total_value
    ).to_dict()
    asset_sectors = risk_pool.set_index("Asset")["Sector"].to_dict()

    for asset, cap_value in asset_caps.items():
        targets[asset] = min(targets[asset], cap_value)

    for sector, sector_assets in risk_pool.groupby("Sector")["Asset"]:
        sector_target = sum(targets[asset] for asset in sector_assets)
        if sector_target > sector_cap_value and sector_target > 0:
            scale_factor = sector_cap_value / sector_target
            for asset in sector_assets:
                targets[asset] *= scale_factor

    proceeds = sum(
        max(0, row["Market Value"] - targets[row["Asset"]])
        for _, row in risk_pool.iterrows()
    )
    remaining_cash = proceeds

    sector_targets = {
        sector: sum(targets[asset] for asset in sector_assets)
        for sector, sector_assets in risk_pool.groupby("Sector")["Asset"]
    }

    candidates = risk_pool.sort_values("Asset Allocation %")
    for _, candidate in candidates.iterrows():
        if remaining_cash <= 0:
            break

        asset = candidate["Asset"]
        sector = candidate["Sector"]
        asset_capacity = max(0, asset_caps[asset] - targets[asset])
        sector_capacity = max(0, sector_cap_value - sector_targets.get(sector, 0))
        buy_value = min(remaining_cash, asset_capacity, sector_capacity)

        if buy_value <= 0:
            continue

        targets[asset] += buy_value
        sector_targets[sector] = sector_targets.get(sector, 0) + buy_value
        remaining_cash -= buy_value

    trades = []
    for _, row in risk_pool.iterrows():
        asset = row["Asset"]
        trade_value = targets[asset] - row["Market Value"]

        if abs(trade_value) < 1 or row["Current Price"] <= 0:
            continue

        action = "BUY" if trade_value > 0 else "SELL"
        shares = math.ceil(abs(trade_value) / row["Current Price"])
        estimated_value = shares * row["Current Price"]
        reason = (
            "Redeploy proceeds into an under-limit asset"
            if action == "BUY"
            else "Reduce asset/sector concentration"
        )

        trades.append({
            "Action": action,
            "Asset": asset,
            "Sector": row["Sector"],
            "Shares": shares,
            "Current Price": row["Current Price"],
            "Estimated Value": estimated_value,
            "Reason": reason,
        })

    if remaining_cash > 1:
        cash_assets = holdings[
            holdings["Sector"].eq("Cash") & (holdings["Current Price"] > 0)
        ].copy()

        if not cash_assets.empty:
            cash_asset = cash_assets.iloc[0]
            shares = math.ceil(remaining_cash / cash_asset["Current Price"])
            trades.append({
                "Action": "BUY",
                "Asset": cash_asset["Asset"],
                "Sector": cash_asset["Sector"],
                "Shares": shares,
                "Current Price": cash_asset["Current Price"],
                "Estimated Value": shares * cash_asset["Current Price"],
                "Reason": "Hold remaining proceeds without adding concentration risk",
            })

    return pd.DataFrame(
        trades,
        columns=[
            "Action",
            "Asset",
            "Sector",
            "Shares",
            "Current Price",
            "Estimated Value",
            "Reason",
        ],
    )


def build_risk_engine(holdings, sector_risk_alloc_pct, total_value, sector_risk_total):
    asset_risk_pool = holdings[~holdings["Sector"].isin(SECTOR_RISK_EXCLUSIONS)].copy()
    asset_risk_pool = add_dynamic_asset_limits(asset_risk_pool)
    asset_violations = asset_risk_pool[
        asset_risk_pool["Asset Allocation %"] > asset_risk_pool["Asset Limit %"]
    ][["Asset", "Sector", "Market Value", "Asset Allocation %", "Asset Limit %"]].copy()
    asset_violations["Excess %"] = (
        asset_violations["Asset Allocation %"] - asset_violations["Asset Limit %"]
    )
    asset_violations = asset_violations[
        [
            "Asset",
            "Sector",
            "Asset Allocation %",
            "Asset Limit %",
            "Excess %",
        ]
    ]

    sector_violations = sector_risk_alloc_pct[
        sector_risk_alloc_pct > SECTOR_WEIGHT_LIMIT
    ].reset_index()
    sector_violations.columns = ["Sector", "Sector Weight %"]
    sector_violations["Sector Value"] = (
        sector_violations["Sector Weight %"] / 100
    ) * total_value
    sector_violations["Limit %"] = SECTOR_WEIGHT_LIMIT
    sector_violations["Excess %"] = (
        sector_violations["Sector Weight %"] - SECTOR_WEIGHT_LIMIT
    )
    sector_violations = sector_violations[
        [
            "Sector",
            "Sector Weight %",
            "Limit %",
            "Excess %",
        ]
    ]

    diversification_score = max(
        0,
        100 - (len(asset_violations) * 3) - (len(sector_violations) * 3),
    )
    risk_score = risk_label_from_score(diversification_score)

    risk_summary = pd.DataFrame({
        "Metric": [
            "Asset Limit",
            "Sector Limit",
            "Single-Asset Cap",
            "Asset Violations",
            "Sector Violations",
            "Diversification Score",
            "Risk Score",
        ],
        "Value": [
            "Sector limit / asset count",
            f"{SECTOR_WEIGHT_LIMIT}%",
            f"{SINGLE_ASSET_WEIGHT_LIMIT}%",
            len(asset_violations),
            len(sector_violations),
            f"{diversification_score} / 100",
            risk_score,
        ],
    })

    return risk_summary, asset_violations, sector_violations


def style_dashboard_sheet(
    worksheet,
    asset_count,
    sector_count,
    risk_summary_count,
    asset_violation_count,
    sector_violation_count,
    suggestion_count,
):
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    title_fill = PatternFill("solid", fgColor=DARK_OLIVE)
    header_fill = PatternFill("solid", fgColor=DARK_OLIVE)
    odd_fill = PatternFill("solid", fgColor=WARM_BEIGE)
    even_fill = PatternFill("solid", fgColor=WARM_WHITE)
    summary_fill = PatternFill("solid", fgColor=SOFT_BEIGE)

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = None
    worksheet.merge_cells("A1:C1")
    worksheet["A1"] = "Portfolio Dashboard"
    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = Font(name="Georgia", color=CREAM, bold=True, size=16)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 25

    widths = {
        "A": 28,
        "B": 18,
        "C": 22,
        "D": 14,
        "E": 28,
        "F": 18,
        "G": 22,
        "H": 42,
        "I": 18,
        "J": 18,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    table_ranges = [
        (SUMMARY_START_ROW + 1, SUMMARY_START_ROW + 1 + 3, 2),
        (ASSET_TABLE_START_ROW + 1, ASSET_TABLE_START_ROW + 1 + asset_count, 3),
        (
            ASSET_TABLE_START_ROW + asset_count + 4,
            ASSET_TABLE_START_ROW + asset_count + 4 + sector_count,
            2,
        ),
        (
            ASSET_TABLE_START_ROW + asset_count + sector_count + 7,
            ASSET_TABLE_START_ROW + asset_count + sector_count + 7 + risk_summary_count,
            2,
        ),
        (
            ASSET_TABLE_START_ROW + asset_count + sector_count + risk_summary_count + 10,
            ASSET_TABLE_START_ROW + asset_count + sector_count + risk_summary_count + 10 + asset_violation_count,
            5,
        ),
        (
            ASSET_TABLE_START_ROW + asset_count + sector_count + risk_summary_count + asset_violation_count + 13,
            ASSET_TABLE_START_ROW + asset_count + sector_count + risk_summary_count + asset_violation_count + 13 + sector_violation_count,
            4,
        ),
        (
            ASSET_TABLE_START_ROW + asset_count + sector_count + risk_summary_count + asset_violation_count + sector_violation_count + 16,
            ASSET_TABLE_START_ROW + asset_count + sector_count + risk_summary_count + asset_violation_count + sector_violation_count + 16 + suggestion_count,
            7,
        ),
    ]

    for start_row, end_row, end_col in table_ranges:
        for row in range(start_row, end_row + 1):
            is_header = row == start_row
            fill = header_fill if is_header else odd_fill if row % 2 == 0 else even_fill
            font = (
                Font(name="Georgia", color=CREAM, bold=True)
                if is_header
                else Font(name="Georgia", color=TEXT_DARK)
            )

            for col in range(1, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = fill
                cell.font = font
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="left" if col == 1 else "center",
                    vertical="center",
                    wrap_text=col >= 7,
                )

    for row in range(SUMMARY_START_ROW + 2, SUMMARY_START_ROW + 5):
        for col in range(1, 3):
            worksheet.cell(row=row, column=col).fill = summary_fill

    for row in range(ASSET_TABLE_START_ROW + 2, ASSET_TABLE_START_ROW + 2 + asset_count):
        worksheet.cell(row=row, column=2).number_format = '#,##0.00'
        worksheet.cell(row=row, column=3).number_format = '0.00"%"'

    sector_header_row = ASSET_TABLE_START_ROW + asset_count + 4
    for row in range(sector_header_row + 1, sector_header_row + 1 + sector_count):
        worksheet.cell(row=row, column=2).number_format = '0.00"%"'

    risk_summary_header_row = ASSET_TABLE_START_ROW + asset_count + sector_count + 7
    asset_risk_header_row = risk_summary_header_row + risk_summary_count + 3
    sector_risk_header_row = asset_risk_header_row + asset_violation_count + 3
    suggestion_header_row = sector_risk_header_row + sector_violation_count + 3
    suggestion_title_row = suggestion_header_row - 1

    worksheet.merge_cells(
        start_row=suggestion_title_row,
        start_column=1,
        end_row=suggestion_title_row,
        end_column=7,
    )
    suggestion_title = worksheet.cell(row=suggestion_title_row, column=1)
    suggestion_title.value = "Holistic Rebalance Suggestion"
    suggestion_title.fill = title_fill
    suggestion_title.font = Font(name="Georgia", color=CREAM, bold=True, size=12)
    suggestion_title.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[suggestion_title_row].height = 22

    for row in range(asset_risk_header_row + 1, asset_risk_header_row + 1 + asset_violation_count):
        worksheet.cell(row=row, column=3).number_format = '0.00"%"'
        worksheet.cell(row=row, column=4).number_format = '0.00"%"'
        worksheet.cell(row=row, column=5).number_format = '0.00"%"'

    for row in range(sector_risk_header_row + 1, sector_risk_header_row + 1 + sector_violation_count):
        worksheet.cell(row=row, column=2).number_format = '0.00"%"'
        worksheet.cell(row=row, column=3).number_format = '0.00"%"'
        worksheet.cell(row=row, column=4).number_format = '0.00"%"'

    for row in range(suggestion_header_row + 1, suggestion_header_row + 1 + suggestion_count):
        worksheet.cell(row=row, column=4).number_format = '#,##0'
        worksheet.cell(row=row, column=5).number_format = '#,##0.00'
        worksheet.cell(row=row, column=6).number_format = '#,##0.00'

    for row in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 18
    worksheet.row_dimensions[1].height = 25

    for row in range(suggestion_header_row + 1, suggestion_header_row + suggestion_count + 1):
        worksheet.row_dimensions[row].height = 36


def style_transactions_sheet(worksheet):
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    header_fill = PatternFill("solid", fgColor=DARK_OLIVE)
    odd_fill = PatternFill("solid", fgColor=WARM_BEIGE)
    even_fill = PatternFill("solid", fgColor=WARM_WHITE)

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    widths = {
        "A": 20,
        "B": 28,
        "C": 12,
        "D": 14,
        "E": 14,
        "F": 18,
        "G": 12,
        "H": 20,
        "I": 16,
        "J": 16,
        "K": 16,
        "L": 22,
        "M": 18,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in range(1, worksheet.max_row + 1):
        is_header = row == 1
        fill = header_fill if is_header else odd_fill if row % 2 == 0 else even_fill
        font = (
            Font(name="Georgia", color=CREAM, bold=True)
            if is_header
            else Font(name="Georgia", color=TEXT_DARK)
        )

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="left" if col in [1, 2, 3, 6, 8] else "center",
                vertical="center",
                wrap_text=True,
            )

    for row in range(2, worksheet.max_row + 1):
        for col in [4, 5, 7, 9, 10, 11, 12, 13]:
            worksheet.cell(row=row, column=col).number_format = '#,##0.00'


def style_nse_prices_sheet(worksheet):
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    header_fill = PatternFill("solid", fgColor=DARK_OLIVE)
    odd_fill = PatternFill("solid", fgColor=WARM_BEIGE)
    even_fill = PatternFill("solid", fgColor=WARM_WHITE)

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    widths = {
        "A": 28,
        "B": 18,
        "C": 18,
        "D": 24,
        "E": 22,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in range(1, worksheet.max_row + 1):
        is_header = row == 1
        fill = header_fill if is_header else odd_fill if row % 2 == 0 else even_fill
        font = (
            Font(name="Georgia", color=CREAM, bold=True)
            if is_header
            else Font(name="Georgia", color=TEXT_DARK)
        )

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="left" if col == 1 else "center",
                vertical="center",
                wrap_text=True,
            )

    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=3).number_format = '#,##0.00'


def normalize_asset_name(name):
    if pd.isna(name):
        return ""
    normalized = str(name).lower()
    normalized = re.sub(r"[^a-z0-9\\s]", "", normalized)
    return normalized.strip()


def find_matching_holding_asset(dividend_asset, holdings):
    normalized_div = normalize_asset_name(dividend_asset)
    if not normalized_div:
        return None

    for asset in holdings["Asset"].astype(str).unique():
        normalized_hold = normalize_asset_name(asset)
        if normalized_div == normalized_hold:
            return asset
        if normalized_div in normalized_hold or normalized_hold in normalized_div:
            return asset

    return None


def load_dividend_sheet(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name=DIVIDEND_INPUT_SHEET)
    except Exception:
        return pd.DataFrame()

    df.columns = df.columns.astype(str).str.strip()
    return df


def build_period_returns(history_df, freq, label):
    if history_df.empty:
        return pd.DataFrame(columns=[label, "Return %"])

    history = history_df.copy()
    history["Date"] = pd.to_datetime(history["Date"], errors="coerce")
    history = history.dropna(subset=["Date"]).sort_values("Date")
    history = history.set_index("Date")
    values = pd.to_numeric(history["Portfolio Value"], errors="coerce").fillna(0)
    if values.empty:
        return pd.DataFrame(columns=[label, "Return %"])

    if freq == "M":
        resample_freq = "ME"
    elif freq == "Q":
        resample_freq = "QE"
    else:
        resample_freq = "YE"

    period_data = values.resample(resample_freq).agg(["first", "last"]).dropna()
    period_data = period_data[period_data["first"] > 0]
    period_data["Return %"] = ((period_data["last"] - period_data["first"]) / period_data["first"]) * 100

    if freq == "M":
        labels = period_data.index.to_period("M").astype(str)
    elif freq == "Q":
        labels = period_data.index.to_period("Q").astype(str)
    else:
        labels = period_data.index.to_period("Y").astype(str)

    return pd.DataFrame({label: labels, "Return %": period_data["Return %"].round(2).fillna(0)})


def build_performance_history(holdings, output_file, current_date, current_value):
    history = read_previous_sheet(output_file, PERFORMANCE_HISTORY_SHEET, ["Date", "Portfolio Value", "Largest Asset", "Largest Sector"])
    if not history.empty:
        history["Date"] = pd.to_datetime(history["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    else:
        history = pd.DataFrame(columns=["Date", "Portfolio Value", "Largest Asset", "Largest Sector"])

    # Find largest asset and sector
    largest_asset = "N/A"
    largest_asset_value = 0
    largest_sector = "N/A"
    largest_sector_value = 0
    
    if not holdings.empty:
        # Find largest asset by market value
        asset_col = next((col for col in holdings.columns if col.lower() == "asset"), None)
        market_col = next((col for col in holdings.columns if col.lower() == "market value"), None)
        sector_col = next((col for col in holdings.columns if col.lower() == "sector"), None)
        
        if asset_col and market_col:
            holdings_copy = holdings.copy()
            holdings_copy[market_col] = pd.to_numeric(holdings_copy[market_col], errors="coerce")
            max_asset_idx = holdings_copy[market_col].idxmax()
            if pd.notna(max_asset_idx):
                largest_asset = holdings_copy.loc[max_asset_idx, asset_col]
                largest_asset_value = holdings_copy.loc[max_asset_idx, market_col]
        
        # Find largest sector by total market value
        if sector_col and market_col:
            sector_totals = holdings_copy.groupby(sector_col)[market_col].sum()
            if not sector_totals.empty:
                largest_sector_idx = sector_totals.idxmax()
                largest_sector = largest_sector_idx
                largest_sector_value = sector_totals.loc[largest_sector_idx]
    
    latest_entry = pd.DataFrame([{
        "Date": current_date, 
        "Portfolio Value": current_value,
        "Largest Asset": f"{largest_asset} (KES {largest_asset_value:,.0f})",
        "Largest Sector": f"{largest_sector} (KES {largest_sector_value:,.0f})"
    }])
    history = pd.concat([history, latest_entry], ignore_index=True)
    history["Date"] = pd.to_datetime(history["Date"], errors="coerce")
    history = history.dropna(subset=["Date"])
    history["Date"] = history["Date"].dt.strftime("%Y-%m-%d")
    history = history.drop_duplicates(subset=["Date"], keep="last").sort_values("Date")

    monthly_returns = build_period_returns(history, "M", "Month")
    quarterly_returns = build_period_returns(history, "Q", "Quarter")
    yearly_returns = build_period_returns(history, "Y", "Year")

    return history, monthly_returns, quarterly_returns, yearly_returns


def build_dividend_table(holdings, dividend_sheet):
    if dividend_sheet.empty:
        return pd.DataFrame(), pd.DataFrame()

    dividends = dividend_sheet.copy()
    for col in ["Dividend per Share", "Shares", "Total Dividend"]:
        if col not in dividends.columns:
            dividends[col] = None

    dividends["Dividend per Share"] = pd.to_numeric(dividends["Dividend per Share"], errors="coerce")
    dividends["Shares"] = pd.to_numeric(dividends["Shares"], errors="coerce").fillna(0)
    dividends["Total Dividend"] = pd.to_numeric(dividends["Total Dividend"], errors="coerce")

    dividends["Annual Dividend"] = dividends["Total Dividend"].copy()
    computed_total = dividends["Dividend per Share"] * dividends["Shares"]
    use_computed = dividends["Annual Dividend"].isna() | (dividends["Annual Dividend"] == 0)
    dividends.loc[use_computed, "Annual Dividend"] = computed_total.loc[use_computed]
    dividends["Annual Dividend"] = dividends["Annual Dividend"].fillna(0)

    market_values = []
    dividend_yields = []
    for _, row in dividends.iterrows():
        holding_asset = find_matching_holding_asset(row.get("Asset"), holdings)
        if holding_asset is None:
            market_values.append(None)
            dividend_yields.append(0)
            continue

        hold_row = holdings[holdings["Asset"] == holding_asset].iloc[0]
        market_value = pd.to_numeric(hold_row.get("Market Value"), errors="coerce")
        market_values.append(market_value)
        if market_value and market_value > 0:
            dividend_yields.append((row["Annual Dividend"] / market_value) * 100)
        else:
            dividend_yields.append(0)

    dividends["Market Value"] = pd.to_numeric(market_values, errors="coerce")
    dividends["Dividend Yield %"] = pd.to_numeric(pd.Series(dividend_yields), errors="coerce").fillna(0)

    total_dividend = dividends["Annual Dividend"].sum()
    total_market_value = holdings["Market Value"].sum()
    overall_yield = (total_dividend / total_market_value * 100) if total_market_value else 0

    summary = pd.DataFrame({
        "Metric": ["Total Annual Dividend", "Portfolio Dividend Yield %"],
        "Value": [total_dividend, overall_yield],
    })

    return dividends, summary


def style_dividends_sheet(worksheet):
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    header_fill = PatternFill("solid", fgColor=DARK_OLIVE)
    odd_fill = PatternFill("solid", fgColor=WARM_BEIGE)
    even_fill = PatternFill("solid", fgColor=WARM_WHITE)

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    widths = {
        "A": 28,
        "B": 18,
        "C": 18,
        "D": 12,
        "E": 18,
        "F": 18,
        "G": 18,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in range(1, worksheet.max_row + 1):
        is_header = row == 1
        fill = header_fill if is_header else odd_fill if row % 2 == 0 else even_fill
        font = (
            Font(name="Georgia", color=CREAM, bold=True)
            if is_header
            else Font(name="Georgia", color=TEXT_DARK)
        )

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="left" if col == 1 else "center",
                vertical="center",
                wrap_text=True,
            )

    for row in range(2, worksheet.max_row + 1):
        for col in [3, 4, 5, 6, 7]:
            try:
                worksheet.cell(row=row, column=col).number_format = '#,##0.00'
            except Exception:
                pass


def style_performance_history_sheet(worksheet):
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    header_fill = PatternFill("solid", fgColor=DARK_OLIVE)
    odd_fill = PatternFill("solid", fgColor=WARM_BEIGE)
    even_fill = PatternFill("solid", fgColor=WARM_WHITE)

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    widths = {
        "A": 18,
        "B": 18,
        "C": 16,
        "D": 16,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in range(1, worksheet.max_row + 1):
        is_header = row == 1
        fill = header_fill if is_header else odd_fill if row % 2 == 0 else even_fill
        font = (
            Font(name="Georgia", color=CREAM, bold=True)
            if is_header
            else Font(name="Georgia", color=TEXT_DARK)
        )

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="left" if col == 1 else "center",
                vertical="center",
                wrap_text=True,
            )

    for row in range(2, worksheet.max_row + 1):
        for col in [2, 3, 4]:
            try:
                worksheet.cell(row=row, column=col).number_format = '#,##0.00'
            except Exception:
                pass


def add_performance_history_chart(worksheet, history_rows):
    chart = LineChart()
    chart.title = "Portfolio Growth"
    chart.y_axis.title = "Portfolio Value"
    chart.x_axis.title = "Date"
    chart.height = 10
    chart.width = 20

    if history_rows < 2:
        return

    data = Reference(worksheet, min_col=2, min_row=1, max_row=history_rows + 1)
    dates = Reference(worksheet, min_col=1, min_row=2, max_row=history_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(dates)
    worksheet.add_chart(chart, "F2")


def style_performance_history_sheet(worksheet):
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    header_fill = PatternFill("solid", fgColor=DARK_OLIVE)
    odd_fill = PatternFill("solid", fgColor=WARM_BEIGE)
    even_fill = PatternFill("solid", fgColor=WARM_WHITE)

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    widths = {
        "A": 20,
        "B": 18,
        "C": 18,
        "D": 18,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in range(1, worksheet.max_row + 1):
        is_header = row == 1
        fill = header_fill if is_header else odd_fill if row % 2 == 0 else even_fill
        font = (
            Font(name="Georgia", color=CREAM, bold=True)
            if is_header
            else Font(name="Georgia", color=TEXT_DARK)
        )

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="left" if col == 1 else "center",
                vertical="center",
                wrap_text=True,
            )

    for row in range(2, worksheet.max_row + 1):
        for col in [2, 3, 4]:
            try:
                worksheet.cell(row=row, column=col).number_format = '#,##0.00'
            except Exception:
                pass


def style_holdings_sheet(worksheet):
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    header_fill = PatternFill("solid", fgColor=DARK_OLIVE)
    odd_fill = PatternFill("solid", fgColor=WARM_BEIGE)
    even_fill = PatternFill("solid", fgColor=WARM_WHITE)

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    # sensible defaults for common holdings columns
    widths = {
        "A": 28,
        "B": 18,
        "C": 12,
        "D": 14,
        "E": 14,
        "F": 18,
        "G": 16,
        "H": 16,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in range(1, worksheet.max_row + 1):
        is_header = row == 1
        fill = header_fill if is_header else odd_fill if row % 2 == 0 else even_fill
        font = (
            Font(name="Georgia", color=CREAM, bold=True)
            if is_header
            else Font(name="Georgia", color=TEXT_DARK)
        )

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="left" if col == 1 else "center",
                vertical="center",
                wrap_text=True,
            )

    # Apply number formats based on header names
    headers = [
        worksheet.cell(row=1, column=col).value
        for col in range(1, worksheet.max_column + 1)
    ]
    header_to_format = {
        "Shares": '#,##0',
        "Buy Price": '#,##0.00',
        "Current Price": '#,##0.00',
        "Market Value": '#,##0.00',
        "Gain/Loss": '#,##0.00',
        "Asset Allocation %": '0.00"%"',
        "Asset Allocation%": '0.00"%"',
        "Average Return %": '0.00"%"',
    }

    for idx, h in enumerate(headers, start=1):
        if h in header_to_format:
            for row in range(2, worksheet.max_row + 1):
                try:
                    worksheet.cell(row=row, column=idx).number_format = header_to_format[h]
                except Exception:
                    pass


def add_dashboard_charts(
    file_path,
    asset_count,
    sector_count,
    risk_summary_count,
    asset_violation_count,
    sector_violation_count,
    suggestion_count,
    performance_history_count,
):
    try:
        workbook = load_workbook(file_path)
    except Exception as e:
        print(f"Warning: could not open '{file_path}' to add charts: {e}")
        return
    worksheet = workbook["Dashboard1"]
    style_dashboard_sheet(
        worksheet,
        asset_count,
        sector_count,
        risk_summary_count,
        asset_violation_count,
        sector_violation_count,
        suggestion_count,
    )

    asset_header_row = ASSET_TABLE_START_ROW + 1
    asset_first_row = asset_header_row + 1
    asset_last_row = asset_header_row + asset_count

    sector_header_row = ASSET_TABLE_START_ROW + asset_count + 4
    sector_first_row = sector_header_row + 1
    sector_last_row = sector_header_row + sector_count

    if asset_count:
        pie_chart = PieChart()
        pie_chart.title = "Asset Allocation"
        pie_chart.height = 9
        pie_chart.width = 13

        pie_data = Reference(
            worksheet,
            min_col=3,
            min_row=asset_header_row,
            max_row=asset_last_row,
        )
        pie_labels = Reference(
            worksheet,
            min_col=1,
            min_row=asset_first_row,
            max_row=asset_last_row,
        )

        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_labels)
        worksheet.add_chart(pie_chart, "E2")

    if sector_count:
        bar_chart = BarChart()
        bar_chart.title = "Sector Allocation"
        bar_chart.y_axis.title = "Allocation %"
        bar_chart.x_axis.title = "Sector"
        bar_chart.height = 9
        bar_chart.width = 13

        bar_data = Reference(
            worksheet,
            min_col=2,
            min_row=sector_header_row,
            max_row=sector_last_row,
        )
        bar_categories = Reference(
            worksheet,
            min_col=1,
            min_row=sector_first_row,
            max_row=sector_last_row,
        )

        bar_chart.add_data(bar_data, titles_from_data=True)
        bar_chart.set_categories(bar_categories)
        worksheet.add_chart(bar_chart, "E20")

    if TRANSACTIONS_SHEET in workbook.sheetnames:
        style_transactions_sheet(workbook[TRANSACTIONS_SHEET])

    if "NSE_Prices" in workbook.sheetnames:
        style_nse_prices_sheet(workbook["NSE_Prices"])

    if DIVIDEND_OUTPUT_SHEET in workbook.sheetnames:
        style_dividends_sheet(workbook[DIVIDEND_OUTPUT_SHEET])

    if PERFORMANCE_HISTORY_SHEET in workbook.sheetnames:
        style_performance_history_sheet(workbook[PERFORMANCE_HISTORY_SHEET])
        add_performance_history_chart(workbook[PERFORMANCE_HISTORY_SHEET], performance_history_count)

    if "Holdings" in workbook.sheetnames:
        style_holdings_sheet(workbook["Holdings"])

    if STATE_SHEET in workbook.sheetnames:
        workbook[STATE_SHEET].sheet_state = "hidden"

    workbook.save(file_path)


def add_dashboard_charts_to_workbook(
    workbook,
    asset_count,
    sector_count,
    risk_summary_count,
    asset_violation_count,
    sector_violation_count,
    suggestion_count,
    performance_history_count,
):
    worksheet = workbook["Dashboard1"]
    style_dashboard_sheet(
        worksheet,
        asset_count,
        sector_count,
        risk_summary_count,
        asset_violation_count,
        sector_violation_count,
        suggestion_count,
    )

    asset_header_row = ASSET_TABLE_START_ROW + 1
    asset_first_row = asset_header_row + 1
    asset_last_row = asset_header_row + asset_count

    sector_header_row = ASSET_TABLE_START_ROW + asset_count + 4
    sector_first_row = sector_header_row + 1
    sector_last_row = sector_header_row + sector_count

    if asset_count:
        pie_chart = PieChart()
        pie_chart.title = "Asset Allocation"
        pie_chart.height = 9
        pie_chart.width = 13

        pie_data = Reference(
            worksheet,
            min_col=3,
            min_row=asset_header_row,
            max_row=asset_last_row,
        )
        pie_labels = Reference(
            worksheet,
            min_col=1,
            min_row=asset_first_row,
            max_row=asset_last_row,
        )

        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_labels)
        worksheet.add_chart(pie_chart, "E2")

    if sector_count:
        bar_chart = BarChart()
        bar_chart.title = "Sector Allocation"
        bar_chart.y_axis.title = "Allocation %"
        bar_chart.x_axis.title = "Sector"
        bar_chart.height = 9
        bar_chart.width = 13

        bar_data = Reference(
            worksheet,
            min_col=2,
            min_row=sector_header_row,
            max_row=sector_last_row,
        )
        bar_categories = Reference(
            worksheet,
            min_col=1,
            min_row=sector_first_row,
            max_row=sector_last_row,
        )

        bar_chart.add_data(bar_data, titles_from_data=True)
        bar_chart.set_categories(bar_categories)
        worksheet.add_chart(bar_chart, "E20")

    if TRANSACTIONS_SHEET in workbook.sheetnames:
        style_transactions_sheet(workbook[TRANSACTIONS_SHEET])

    if "NSE_Prices" in workbook.sheetnames:
        style_nse_prices_sheet(workbook["NSE_Prices"])

    if DIVIDEND_OUTPUT_SHEET in workbook.sheetnames:
        style_dividends_sheet(workbook[DIVIDEND_OUTPUT_SHEET])

    if PERFORMANCE_HISTORY_SHEET in workbook.sheetnames:
        style_performance_history_sheet(workbook[PERFORMANCE_HISTORY_SHEET])
        add_performance_history_chart(workbook[PERFORMANCE_HISTORY_SHEET], performance_history_count)

    if "Holdings" in workbook.sheetnames:
        style_holdings_sheet(workbook["Holdings"])

    if STATE_SHEET in workbook.sheetnames:
        workbook[STATE_SHEET].sheet_state = "hidden"


def generate_pdf_report(portfolio_name, performance_history, asset_violations, sector_violations, total_value, output_pdf_path, pdf_password=None):
    """Generate a PDF report with 6-month portfolio performance and risk violations."""
    if not HAS_REPORTLAB:
        print(f"Warning: reportlab not installed - skipping PDF generation for {portfolio_name}")
        return False

    try:
        doc = SimpleDocTemplate(output_pdf_path, pagesize=PDF_PAGE_SIZE, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(name="CustomTitle", parent=styles["Heading1"], fontSize=PDF_TITLE_SIZE, textColor=colors.HexColor("#3B4436"), spaceAfter=0.3*inch, alignment=TA_CENTER)
        story.append(Paragraph(f"{portfolio_name} - Portfolio Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Report Date
        report_date = datetime.now().strftime("%B %d, %Y")
        date_style = ParagraphStyle(name="DateStyle", parent=styles["Normal"], fontSize=PDF_TEXT_SIZE, alignment=TA_CENTER, textColor=colors.HexColor("#2F332E"))
        story.append(Paragraph(f"Generated: {report_date}", date_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Portfolio Value Summary
        heading_style = ParagraphStyle(name="CustomHeading", parent=styles["Heading2"], fontSize=PDF_HEADING_SIZE, textColor=colors.HexColor("#3B4436"), spaceAfter=0.15*inch)
        story.append(Paragraph("Portfolio Performance (Last 6 Months)", heading_style))
        
        # Current Value
        value_style = ParagraphStyle(name="ValueStyle", parent=styles["Normal"], fontSize=12, textColor=colors.HexColor("#2F332E"), spaceAfter=0.1*inch)
        current_value_text = f"Current Total Portfolio Value: <b>KES {total_value:,.2f}</b>"
        story.append(Paragraph(current_value_text, value_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Generate 6-month performance chart
        if not performance_history.empty:
            history_df = performance_history.copy()
            history_df["Date"] = pd.to_datetime(history_df["Date"], errors="coerce")
            six_months_ago = datetime.now() - timedelta(days=180)
            history_df = history_df[history_df["Date"] >= six_months_ago].sort_values("Date")
            
            if not history_df.empty:
                plt.figure(figsize=(8, 4))
                plt.plot(history_df["Date"], history_df["Portfolio Value"], marker='o', color="#3B4436", linewidth=2, markersize=4)
                plt.title("6-Month Portfolio Value Trend", fontsize=12, color="#3B4436")
                plt.xlabel("Date", fontsize=10, color="#2F332E")
                plt.ylabel("Portfolio Value (KES)", fontsize=10, color="#2F332E")
                plt.grid(True, alpha=0.3)
                plt.xticks(rotation=45, fontsize=9)
                plt.tight_layout()
                
                chart_path = os.path.join(tempfile.gettempdir(), "portfolio_chart.png")
                plt.savefig(chart_path, dpi=100, bbox_inches="tight")
                plt.close()
                
                try:
                    story.append(Image(chart_path, width=6*inch, height=3*inch))
                except Exception as e:
                    print(f"Warning: failed to insert chart image: {e}")
                
                # Add performance history table
                story.append(Spacer(1, 0.15*inch))
                story.append(Paragraph("Performance History Details", value_style))
                perf_data = [["Date", "Portfolio Value", "Largest Asset", "Largest Sector"]]
                for _, row in history_df.iterrows():
                    perf_data.append([
                        str(row.get("Date", "")),
                        f"KES {row.get('Portfolio Value', 0):,.0f}",
                        str(row.get("Largest Asset", "N/A")),
                        str(row.get("Largest Sector", "N/A"))
                    ])
                
                perf_table = Table(perf_data, colWidths=[1.2*inch, 1.5*inch, 1.5*inch, 1.3*inch])
                perf_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B4436")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#F1E9CB")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FDFBF7")),
                    ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#B8AA91")),
                    ("FONTSIZE", (0, 1), (-1, -1), 7),
                ]))
                story.append(perf_table)
        
        story.append(Spacer(1, 0.3*inch))
        
        # Risk Violations Section
        if not asset_violations.empty or not sector_violations.empty:
            story.append(Paragraph("Risk Management Violations", heading_style))
            
            if not asset_violations.empty:
                story.append(Paragraph("<b>Asset Concentration Violations:</b>", value_style))
                violation_data = [["Asset", "Market Value", "Allocation %", "Limit %"]]
                for _, row in asset_violations.iterrows():
                    violation_data.append([
                        str(row.get("Asset", "")),
                        f"KES {row.get('Market Value', 0):,.2f}",
                        f"{row.get('Allocation %', 0):.2f}%",
                        f"{row.get('Single Asset Limit', 0):.2f}%"
                    ])
                
                violation_table = Table(violation_data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 1*inch])
                violation_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B4436")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#F1E9CB")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FDFBF7")),
                    ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#B8AA91")),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                ]))
                story.append(violation_table)
                story.append(Spacer(1, 0.2*inch))
            
            if not sector_violations.empty:
                story.append(Paragraph("<b>Sector Concentration Violations:</b>", value_style))
                sector_data = [["Sector", "Market Value", "Allocation %", "Limit %"]]
                for _, row in sector_violations.iterrows():
                    sector_data.append([
                        str(row.get("Sector", "")),
                        f"KES {row.get('Market Value', 0):,.2f}",
                        f"{row.get('Allocation %', 0):.2f}%",
                        f"{row.get('Sector Limit', 0):.2f}%"
                    ])
                
                sector_table = Table(sector_data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 1*inch])
                sector_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B4436")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#F1E9CB")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FDFBF7")),
                    ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#B8AA91")),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                ]))
                story.append(sector_table)
        else:
            story.append(Paragraph("Risk Status: <b>No violations detected.</b>", value_style))
        
        doc.build(story)
        print(f"PDF report generated: {output_pdf_path}")
        
        # Encrypt PDF with password if provided and PyPDF2 is available
        if pdf_password and HAS_PYPDF2:
            try:
                from PyPDF2 import PdfReader
                writer = PdfWriter()
                with open(output_pdf_path, "rb") as input_pdf:
                    reader = PdfReader(input_pdf)
                    for page in reader.pages:
                        writer.add_page(page)
                writer.encrypt(pdf_password)
                with open(output_pdf_path, "wb") as output_pdf:
                    writer.write(output_pdf)
                print(f"PDF encrypted with password protection")
            except Exception as e:
                print(f"Warning: failed to encrypt PDF: {e}")
        elif pdf_password and not HAS_PYPDF2:
            print("Warning: PyPDF2 not installed - PDF will not be password protected")
        
        return True
    except Exception as e:
        print(f"Error generating PDF: {e}")
        import traceback
        traceback.print_exc()
        return False


def send_email_report(recipient_email, pdf_path, portfolio_name, pdf_password=None):
    """Send the generated PDF report via email."""
    if not HAS_EMAIL:
        print("Warning: email libraries not available - skipping email send")
        return False

    if not EMAIL_SMTP_USERNAME or not EMAIL_SMTP_PASSWORD or not EMAIL_FROM:
        print("SMTP email configuration is incomplete. Set EMAIL_SMTP_USERNAME, EMAIL_SMTP_PASSWORD, and optionally EMAIL_FROM.")
        return False

    try:
        message = MIMEMultipart()
        message["From"] = EMAIL_FROM
        message["To"] = recipient_email
        message["Subject"] = f"{portfolio_name} Portfolio Report"

        body = f"Hello,\n\nPlease find attached the latest portfolio report for {portfolio_name}."
        
        if pdf_password:
            body += f"\n\nIMPORTANT - PDF PASSWORD REQUIRED:\nThe PDF is password protected. To open it, enter the following password when prompted:\n\nPassword: {pdf_password}\n\nThis password is based on the first 8 characters of the portfolio filename."
        
        body += "\n\nRegards,\nPortfolio Tracker"
        message.attach(MIMEText(body, "plain"))

        if pdf_path and os.path.exists(pdf_path):
            with open(pdf_path, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename={os.path.basename(pdf_path)}",
            )
            message.attach(part)
        else:
            print(f"Warning: PDF attachment not found at {pdf_path}")

        if EMAIL_USE_SSL:
            server = smtplib.SMTP_SSL(EMAIL_SMTP_SERVER, EMAIL_SMTP_PORT)
        else:
            server = smtplib.SMTP(EMAIL_SMTP_SERVER, EMAIL_SMTP_PORT)
            server.ehlo()
            if EMAIL_USE_TLS:
                server.starttls()
                server.ehlo()

        server.login(EMAIL_SMTP_USERNAME, EMAIL_SMTP_PASSWORD)
        server.sendmail(EMAIL_FROM, recipient_email, message.as_string())
        server.quit()

        print(f"Email sent to {recipient_email}")
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False


# -----------------------------
# Process multiple client portfolio files
# -----------------------------
CLIENTS_DIR = "clients"
REPORTS_DIR = "reports"

def process_portfolio(input_file, output_file):
    # create a backup of the input
    backup_path = f"backup_{os.path.splitext(os.path.basename(input_file))[0]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy(input_file, backup_path)
    print(f"Backup created: {backup_path}")

    # -----------------------------
    # READ RAW EXCEL
    # -----------------------------
    holdings_raw = pd.read_excel(input_file, sheet_name="Holdings", header=None)

    # Find header row dynamically
    header_idx = None

    for i in range(len(holdings_raw)):
        row = holdings_raw.iloc[i].astype(str).str.lower()
        if "asset" in row.values and "sector" in row.values:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Could not find header row in Holdings sheet")

    # Re-read properly
    holdings = pd.read_excel(input_file, sheet_name="Holdings", header=header_idx)
    holdings.columns = holdings.columns.astype(str).str.strip()

    # -----------------------------
    # CRITICAL FIX 1: FORCE NUMERIC CLEANING
    # -----------------------------
    for col in ["Shares", "Buy Price", "Current Price"]:
        if col not in holdings.columns:
            holdings[col] = 0

        numeric_values = pd.to_numeric(holdings[col], errors="coerce")

        if col == "Current Price":
            formula_values = formula_backed_column_values(
                input_file,
                "Holdings",
                header_idx,
                holdings.columns.get_loc(col) + 1,
                len(holdings),
            )
            holdings[col] = numeric_values.fillna(formula_values).fillna(0)
        else:
            holdings[col] = numeric_values.fillna(0)

    # Gain/Loss safety
    if "Gain/Loss" in holdings.columns:
        holdings["Gain/Loss"] = pd.to_numeric(holdings["Gain/Loss"], errors="coerce").fillna(0)
    else:
        holdings["Gain/Loss"] = 0

    current_sector = None
    for row_index, asset_name in holdings["Asset"].items():
        if asset_name in SECTOR_LABELS:
            current_sector = SECTOR_LABELS[asset_name]
        elif pd.notna(asset_name) and pd.isna(holdings.at[row_index, "Sector"]):
            holdings.at[row_index, "Sector"] = current_sector

    # -----------------------------
    # CLEAN DATA (safe filtering)
    # -----------------------------
    invalid_labels = list(SECTOR_LABELS.keys()) + [
        "Sector Value", "Cash", "Total Portfolio Value"
    ]

    holdings = holdings[holdings["Asset"].notna()]
    holdings = holdings[~holdings["Asset"].isin(invalid_labels)]
    holdings = holdings.copy()

    # -----------------------------
    # CRITICAL FIX 2: APPLY LIVE NSE PRICES BEFORE VALUE CALCULATIONS
    # -----------------------------
    print("\nFetching live NSE prices for mapped holdings...")
    holdings = inject_nse_live_prices(holdings)

    holdings["Market Value"] = holdings["Shares"] * holdings["Current Price"]

    print("\n--- CLEAN HOLDINGS SAMPLE ---")
    print(holdings[["Asset", "Shares", "Current Price", "Market Value", "Price Source"]].head(15))

    # -----------------------------
    # TOTAL VALUE
    # -----------------------------
    total_value = holdings["Market Value"].sum()

    print("\nTOTAL PORTFOLIO VALUE:", total_value)

    # Avoid division by zero crash
    if total_value == 0:
        holdings["Asset Allocation %"] = 0
    else:
        holdings["Asset Allocation %"] = (holdings["Market Value"] / total_value) * 100

    print("\n--- CLEAN ASSET ALLOCATION ---")
    print(holdings[["Asset", "Market Value", "Asset Allocation %"]])

    # -----------------------------
    # RETURNS: Average return per asset (percentage)
    # -----------------------------
    buy_prices = pd.to_numeric(holdings.get("Buy Price", 0), errors="coerce").fillna(0)
    current_prices = pd.to_numeric(holdings.get("Current Price", 0), errors="coerce").fillna(0)
    returns = pd.Series(0.0, index=holdings.index)
    nonzero_buy = buy_prices != 0
    returns.loc[nonzero_buy] = ((current_prices.loc[nonzero_buy] - buy_prices.loc[nonzero_buy]) / buy_prices.loc[nonzero_buy]) * 100
    holdings["Average Return %"] = returns.fillna(0)

    # -----------------------------
    # SECTOR ALLOCATION
    # -----------------------------
    if "Sector" not in holdings.columns:
        holdings["Sector"] = "Unknown"

    sector_alloc = holdings.groupby("Sector")["Market Value"].sum()

    if total_value == 0:
        sector_alloc_pct = sector_alloc * 0
    else:
        sector_alloc_pct = (sector_alloc / total_value) * 100

    print("\n--- CLEAN SECTOR ALLOCATION ---")
    print(sector_alloc_pct)

    # -----------------------------
    # PHASE 2 RISK ENGINE
    # -----------------------------
    sector_risk_pool = holdings[~holdings["Sector"].isin(SECTOR_RISK_EXCLUSIONS)].copy()
    sector_risk_alloc = sector_risk_pool.groupby("Sector")["Market Value"].sum()
    sector_risk_total = sector_risk_alloc.sum()

    if sector_risk_total == 0:
        sector_risk_alloc_pct = sector_risk_alloc * 0
    else:
        sector_risk_alloc_pct = (sector_risk_alloc / total_value) * 100

    risk_summary, asset_violations, sector_violations = build_risk_engine(
        holdings,
        sector_risk_alloc_pct,
        total_value,
        sector_risk_total,
    )
    rebalance_plan = build_rebalance_plan(holdings, total_value, sector_risk_total)
    current_state = build_current_state(holdings)

    # Load dividend tracking data and calculate annual dividends and yield
    dividend_sheet = load_dividend_sheet(input_file)
    dividend_table, dividend_summary = build_dividend_table(holdings, dividend_sheet)

    current_date = datetime.now().strftime("%Y-%m-%d")
    performance_history, monthly_returns, quarterly_returns, yearly_returns = build_performance_history(
        holdings,
        output_file,
        current_date,
        total_value,
    )

    previous_state = read_previous_sheet(output_file, STATE_SHEET, STATE_COLUMNS)
    existing_transactions = read_previous_sheet(
        output_file,
        TRANSACTIONS_SHEET,
        TRANSACTION_COLUMNS,
    )
    new_transactions = detect_share_transactions(previous_state, current_state)
    opening_state = current_state
    opening_transactions = (
        build_opening_transactions(opening_state)
        if existing_transactions.empty
        else pd.DataFrame(columns=TRANSACTION_BASE_COLUMNS)
    )
    preserved_transactions = (
        pd.DataFrame(columns=TRANSACTION_BASE_COLUMNS)
        if existing_transactions.empty
        else existing_transactions[TRANSACTION_BASE_COLUMNS]
    )
    transactions = pd.concat(
        [
            opening_transactions,
            preserved_transactions,
            new_transactions,
        ],
        ignore_index=True,
    )
    transactions = calculate_transaction_metrics(transactions, current_state)

    print("\n--- RISK CHECK (DYNAMIC ASSET RULE) ---")

    if asset_violations.empty:
        print("No asset concentration violations.")
    else:
        print(asset_violations)

    print(f"\n--- RISK CHECK ({SECTOR_WEIGHT_LIMIT}% SECTOR RULE) ---")

    if sector_violations.empty:
        print("No sector concentration violations.")
    else:
        print(sector_violations)

    print("\n--- RISK SCORE ---")
    print(risk_summary)

    print("\n--- HOLISTIC REBALANCE SUGGESTION ---")
    if rebalance_plan.empty:
        print("No trades needed.")
    else:
        print(rebalance_plan)

    print("\n--- TRANSACTION ENGINE ---")
    if not opening_transactions.empty:
        print(f"Opening balances recorded: {len(opening_transactions)}")
    elif new_transactions.empty:
        print("No new share changes detected.")
    else:
        print(new_transactions)

    # -----------------------------
    # PLOTS (SAFE GUARDS FIX)
    # -----------------------------
    plt.figure()

    plot_data = holdings.copy()
    plot_data = plot_data.dropna(subset=["Market Value", "Asset Allocation %"])
    plot_data = plot_data[plot_data["Market Value"] > 0]

    if not plot_data.empty:
        plot_data.set_index("Asset")["Asset Allocation %"].plot(kind="pie", autopct="%1.1f%%")
    else:
        print("No valid data for pie chart.")

    plt.ylabel("")
    plt.show()

    plt.figure()

    if not sector_alloc_pct.dropna().empty:
        sector_alloc_pct.plot(kind="bar")
        plt.title("Sector Allocation")
        plt.ylabel("Percentage")
        plt.xticks(rotation=45)
        plt.tight_layout()
    else:
        print("No valid sector data to plot.")

    plt.show()

    # -----------------------------
    # DASHBOARD EXPORT (FIXED)
    # -----------------------------
    summary_df = pd.DataFrame({
        "Metric": [
            "Total Portfolio Value",
            "Number of Assets",
            "Number of Sectors"
        ],
        "Value": [
            total_value,
            len(holdings),
            holdings["Sector"].nunique()
        ]
    })

    asset_table = holdings[["Asset", "Market Value", "Asset Allocation %"]]
    sector_table = sector_alloc_pct.reset_index()
    sector_table.columns = ["Sector", "Allocation %"]
    risk_summary_startrow = ASSET_TABLE_START_ROW + len(asset_table) + len(sector_table) + 6
    asset_violations_startrow = risk_summary_startrow + len(risk_summary) + 3
    sector_violations_startrow = asset_violations_startrow + len(sector_violations) + 3
    rebalance_plan_startrow = sector_violations_startrow + len(sector_violations) + 3

    with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
        summary_df.to_excel(
            writer,
            sheet_name="Dashboard1",
            startrow=SUMMARY_START_ROW,
            index=False,
        )
        asset_table.to_excel(
            writer,
            sheet_name="Dashboard1",
            startrow=ASSET_TABLE_START_ROW,
            index=False,
        )
        # Also include the full Holdings table in the output workbook
        try:
            holdings.to_excel(
                writer,
                sheet_name="Holdings",
                index=False,
            )
        except Exception:
            # Fallback: write a minimal holdings view if full export fails
            holdings[["Asset", "Shares", "Current Price", "Market Value"]].to_excel(
                writer,
                sheet_name="Holdings",
                index=False,
            )

        if not dividend_table.empty:
            dividend_table.to_excel(
                writer,
                sheet_name=DIVIDEND_OUTPUT_SHEET,
                index=False,
            )
            dividend_summary.to_excel(
                writer,
                sheet_name=DIVIDEND_OUTPUT_SHEET,
                startrow=len(dividend_table) + 3,
                index=False,
            )

        performance_history.to_excel(
            writer,
            sheet_name=PERFORMANCE_HISTORY_SHEET,
            index=False,
        )
        monthly_returns.to_excel(
            writer,
            sheet_name=PERFORMANCE_HISTORY_SHEET,
            startrow=len(performance_history) + 3,
            index=False,
        )
        quarterly_returns.to_excel(
            writer,
            sheet_name=PERFORMANCE_HISTORY_SHEET,
            startrow=len(performance_history) + len(monthly_returns) + 7,
            index=False,
        )
        yearly_returns.to_excel(
            writer,
            sheet_name=PERFORMANCE_HISTORY_SHEET,
            startrow=len(performance_history) + len(monthly_returns) + len(quarterly_returns) + 11,
            index=False,
        )

        sector_table.to_excel(
            writer,
            sheet_name="Dashboard1",
            startrow=ASSET_TABLE_START_ROW + len(asset_table) + 3,
            index=False,
        )
        risk_summary.to_excel(
            writer,
            sheet_name="Dashboard1",
            startrow=risk_summary_startrow,
            index=False,
        )
        asset_violations.to_excel(
            writer,
            sheet_name="Dashboard1",
            startrow=asset_violations_startrow,
            index=False,
        )
        sector_violations.to_excel(
            writer,
            sheet_name="Dashboard1",
            startrow=sector_violations_startrow,
            index=False,
        )
        rebalance_plan.to_excel(
            writer,
            sheet_name="Dashboard1",
            startrow=rebalance_plan_startrow,
            index=False,
        )
        nse_price_table = build_nse_price_table(holdings)
        nse_price_table.to_excel(
            writer,
            sheet_name="NSE_Prices",
            startrow=0,
            index=False,
        )
        transactions.to_excel(
            writer,
            sheet_name=TRANSACTIONS_SHEET,
            index=False,
        )
        current_state.to_excel(
            writer,
            sheet_name=STATE_SHEET,
            index=False,
        )

        # Insert charts/styles into the in-memory workbook before saving
        try:
            add_dashboard_charts_to_workbook(
                writer.book,
                len(asset_table),
                len(sector_table),
                len(risk_summary),
                len(asset_violations),
                len(sector_violations),
                len(rebalance_plan),
                len(performance_history),
            )
        except Exception as e:
            print(f"Warning: failed to add in-memory charts: {e}")

    # Generate PDF report
    portfolio_name = os.path.splitext(os.path.basename(input_file))[0]
    pdf_path = os.path.join(REPORTS_DIR, f"{portfolio_name}_Report.pdf")
    
    # Extract password from first 8 characters of portfolio filename
    pdf_password = portfolio_name[:8] if len(portfolio_name) >= 8 else portfolio_name
    
    generate_pdf_report(portfolio_name, performance_history, asset_violations, sector_violations, total_value, pdf_path, pdf_password)

    # Prompt for email
    print(f"\nDashboard Generated Successfully for {os.path.basename(input_file)} -> {os.path.basename(output_file)}")
    
    user_email = input(f"\nEnter email address to receive {portfolio_name} report (or press Enter to skip): ").strip()
    if user_email:
        if not EMAIL_SMTP_USERNAME or not EMAIL_SMTP_PASSWORD or not EMAIL_FROM:
            print("Warning: SMTP environment variables are incomplete. Email cannot be sent until configuration is provided.")

        if send_email_report(user_email, pdf_path, portfolio_name, pdf_password):
            print(f"Report sent to {user_email}")
        else:
            print("Email report could not be sent. Check SMTP settings and try again.")
    else:
        print(f"Report available at: {pdf_path}")
        if pdf_password:
            print(f"PDF Password: {pdf_password}")


def _ensure_dirs():
    if not os.path.exists(CLIENTS_DIR):
        os.makedirs(CLIENTS_DIR)
    if not os.path.exists(REPORTS_DIR):
        os.makedirs(REPORTS_DIR)



def main():
    _ensure_dirs()

    client_files = [f for f in os.listdir(CLIENTS_DIR) if f.lower().endswith(".xlsx")]
    if not client_files:
        print(f"No .xlsx files found in '{CLIENTS_DIR}' - please add Portfolio_Tracker_Kenya.xlsx or other portfolios.")
        return

    for cf in client_files:
        input_file = os.path.join(CLIENTS_DIR, cf)
        base = os.path.splitext(cf)[0]
        output_file = os.path.join(REPORTS_DIR, f"{base}_Dashboard_Output.xlsx")
        print(f"\nProcessing portfolio: {input_file}")
        # remove any previous corrupted output to ensure a fresh write
        try:
            if os.path.exists(output_file):
                os.remove(output_file)
        except Exception:
            pass

        try:
            process_portfolio(input_file, output_file)
        except Exception as e:
            import traceback
            print(f"Error processing {input_file}: {e}")
            traceback.print_exc()


if __name__ == "__main__":
    main()
